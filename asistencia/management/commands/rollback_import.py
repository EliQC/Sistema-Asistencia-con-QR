from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from asistencia.models import Estudiante, Grado, Seccion
import os
import csv
import unicodedata
import re
from django.db import models

try:
    import openpyxl
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


def _norm_key(k):
    if not k:
        return ''
    nk = unicodedata.normalize('NFKD', str(k)).encode('ascii', 'ignore').decode('ascii')
    nk = re.sub(r'[^0-9a-zA-Z]+', '_', nk).strip('_').lower()
    return nk


class Command(BaseCommand):
    help = 'Rollback parcial de importación: elimina estudiantes importados con grado/seccion placeholder (Sin Grado / Sin).\nUsa --dry-run para ver qué se eliminaría.'

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Ruta al archivo subido (archivo en media/uploads o ruta absoluta)')
        parser.add_argument('--dry-run', action='store_true', help='No borra, solo muestra qué se eliminaría')
        parser.add_argument('--yes', action='store_true', help='Confirma la eliminación sin pedir interacción')

    def handle(self, *args, **options):
        filepath = options['filepath']
        dry = options.get('dry_run')
        assume_yes = options.get('yes')

        # If only filename provided, assume media/uploads
        if not os.path.isabs(filepath):
            uploads_dir = os.path.join(settings.MEDIA_ROOT or 'media', 'uploads')
            candidate = os.path.join(uploads_dir, filepath)
            if os.path.exists(candidate):
                filepath = candidate

        if not os.path.exists(filepath):
            raise CommandError(f'Archivo no encontrado: {filepath}')

        # parse file to extract DNIs
        dnis = set()
        _, ext = os.path.splitext(filepath)
        if ext.lower() in ('.xls', '.xlsx'):
            if not _HAS_OPENPYXL:
                raise CommandError('openpyxl no está instalado; instala openpyxl o convierte el archivo a CSV')
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [_norm_key(h) for h in raw_headers]
            dni_idx = None
            for i, h in enumerate(headers):
                if 'dni' in h or 'documento' in h:
                    dni_idx = i
                    break
            if dni_idx is None:
                raise CommandError('No se encontró columna DNI en el archivo')
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[dni_idx]
                if val:
                    dnis.add(str(val).strip())
        elif ext.lower() == '.csv':
            with open(filepath, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                # find dni-key
                # normalize keys of first row
                rows = list(reader)
                if not rows:
                    raise CommandError('CSV vacío')
                nk_map = { _norm_key(k): k for k in rows[0].keys() }
                dni_key = None
                for nk in nk_map.keys():
                    if 'dni' in nk or 'documento' in nk:
                        dni_key = nk_map[nk]
                        break
                if not dni_key:
                    raise CommandError('No se encontró columna DNI en el CSV')
                for r in rows:
                    v = r.get(dni_key)
                    if v:
                        dnis.add(str(v).strip())
        else:
            raise CommandError('Formato no soportado. Usa .csv o .xlsx')

        if not dnis:
            self.stdout.write(self.style.WARNING('No se encontraron DNIs en el archivo. Nada para hacer.'))
            return

        # Find students matching DNIs with placeholder grado/seccion
        qs_placeholder = Estudiante.objects.filter(dni__in=list(dnis)).filter(models.Q(grado__nombre='Sin Grado') | models.Q(seccion__nombre='Sin'))
        count = qs_placeholder.count()

        if count == 0:
            self.stdout.write(self.style.WARNING('No se encontraron estudiantes con grado/sección placeholder entre los DNIs del archivo.'))
            return

        self.stdout.write(self.style.WARNING(f'Se encontraron {count} estudiantes candidatos para eliminar (grado=Sin Grado o seccion=Sin).'))
        for e in qs_placeholder.order_by('dni'):
            self.stdout.write(f' - {e.dni} | {e.nombre} {e.apellido} | grado={e.grado.nombre} | seccion={e.seccion.nombre}')

        if dry:
            self.stdout.write(self.style.SUCCESS('Dry-run activado; no se borró nada.'))
            return

        if not assume_yes:
            confirm = input('Confirmar eliminación de estos estudiantes? escriba SI para confirmar: ')
            if confirm.strip().upper() != 'SI':
                self.stdout.write(self.style.ERROR('Operación cancelada por el usuario.'))
                return

        # Perform deletion
        deleted_info = []
        for e in qs_placeholder:
            dni = e.dni
            e.delete()
            deleted_info.append(dni)

        self.stdout.write(self.style.SUCCESS(f'Estudiantes eliminados: {len(deleted_info)}'))
        for d in deleted_info:
            self.stdout.write(f' * {d}')
