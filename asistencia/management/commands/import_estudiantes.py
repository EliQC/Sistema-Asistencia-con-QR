from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from asistencia.models import Estudiante, Apoderado, Grado, Seccion
import os
import csv
import qrcode
from io import BytesIO
import re
import unicodedata
import json
import datetime as _dt

try:
    import openpyxl
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


class Command(BaseCommand):
    help = 'Importa estudiantes desde un archivo Excel (.xlsx) o CSV. Genera QR por DNI automáticamente.'

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Ruta al archivo .xlsx o .csv a importar')
        parser.add_argument('--periodo', type=int, help='Año escolar (periodo) a asignar', default=None)

    def handle(self, *args, **options):
        filepath = options['filepath']
        periodo_override = options.get('periodo')

        if not os.path.exists(filepath):
            raise CommandError(f'El archivo {filepath} no existe')

        filename = os.path.basename(filepath)
        _, ext = os.path.splitext(filename)
        rows = []

        def _norm_key(k):
            if not k:
                return ''
            # remover acentos
            nk = unicodedata.normalize('NFKD', str(k)).encode('ascii', 'ignore').decode('ascii')
            # reemplazar no-alphanum por guion bajo
            nk = re.sub(r'[^0-9a-zA-Z]+', '_', nk).strip('_').lower()
            return nk

        if ext.lower() in ('.xls', '.xlsx'):
            if not _HAS_OPENPYXL:
                raise CommandError('openpyxl no está instalado. Instala con: pip install openpyxl')
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [_norm_key(h) for h in raw_headers]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(len(headers))})
        elif ext.lower() == '.csv':
            with open(filepath, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for r in reader:
                    # normalize keys: quitar acentos/espacios y convertir a snake_case lowercase
                    normalized = {}
                    for k, v in r.items():
                        nk = _norm_key(k)
                        normalized[nk] = v
                    rows.append(normalized)
        else:
            raise CommandError('Formato no soportado. Usa .xlsx o .csv')

        media_qr_dir = os.path.join(settings.MEDIA_ROOT or 'media', 'qrcodes')
        os.makedirs(media_qr_dir, exist_ok=True)

        # If this file was uploaded via the web uploader, a companion
        # status JSON is expected at <filepath>.status.json. We'll update
        # it periodically with progress (processed/total) so the UI can show a
        # progress bar.
        status_path = f"{filepath}.status.json"

        created = 0
        updated = 0
        errors = []
        total_rows = len(rows)
        processed = 0

        def _write_progress(status_obj):
            try:
                with open(status_path, 'w', encoding='utf-8') as sf:
                    json.dump(status_obj, sf, default=str, ensure_ascii=False, indent=2)
            except Exception:
                pass

        # initialize progress if status file exists
        if os.path.exists(status_path):
            _write_progress({'status': 'processing', 'started_at': _dt.datetime.now().isoformat(), 'processed': 0, 'total': total_rows})

        for r in rows:
            # Campos esperados (flexible): nombres, apellido_paterno, apellido_materno, dni, fecha_nacimiento, seccion, apoderado_*
            # Mapear posibles columnas normalizadas según el formato entregado
            nombre = (r.get('nombres') or r.get('nombre') or r.get('first_name') or '')
            apellido_p = (r.get('apellido_paterno') or r.get('apellido_p') or r.get('apellido') or '')
            apellido_m = (r.get('apellido_materno') or r.get('apellido_m') or '')
            apellidos_y_nombres = (r.get('apellidos_y_nombres') or r.get('apellidos_y_nombres') or r.get('apellidos_y_nombres') or '')
            # tolerate common misspelling 'apelidos_y_nombres'
            if not apellidos_y_nombres:
                apellidos_y_nombres = (r.get('apelidos_y_nombres') or '')

            apellido = ''
            if apellido_p or apellido_m:
                apellido = ' '.join([x for x in [apellido_p, apellido_m] if x])
            elif apellidos_y_nombres:
                # intentar dividir "APELLIDOS Y NOMBRES" en apellidos y nombres
                ac = str(apellidos_y_nombres)
                if ',' in ac:
                    parts = [p.strip() for p in ac.split(',', 1)]
                    apellido = parts[0]
                    if not nombre:
                        nombre = parts[1]
                else:
                    words = ac.split()
                    if len(words) >= 3:
                        # tomar los últimos dos como apellidos
                        apellido = ' '.join(words[-2:])
                        if not nombre:
                            nombre = ' '.join(words[:-2])
                    else:
                        # no hay forma clara: asignar todo a apellido
                        apellido = ac

            if nombre:
                nombre = str(nombre).strip()
            if apellido:
                apellido = str(apellido).strip()

            dni = str((r.get('dni') or r.get('documento') or '')).strip()
            fecha_nacimiento = r.get('fecha_nacimiento') or r.get('birthdate') or None
            # Intentar parsear fecha si viene como string o tipo datetime
            if fecha_nacimiento:
                try:
                    # openpyxl may return a date/datetime object
                    if isinstance(fecha_nacimiento, _dt.datetime):
                        fecha_nacimiento = fecha_nacimiento.date()
                    elif isinstance(fecha_nacimiento, _dt.date):
                        # already a date
                        pass
                    else:
                        # intentar parsear string en formatos comunes
                        fs = str(fecha_nacimiento).strip()
                        parsed = None
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                            try:
                                parsed = _dt.datetime.strptime(fs, fmt).date()
                                break
                            except Exception:
                                continue
                        fecha_nacimiento = parsed
                except Exception:
                    fecha_nacimiento = None

            # El archivo tiene SECCION y GRADO; pueden venir en columnas separadas o combinadas
            # Soportamos variantes comunes de encabezados: 'grado', 'grade', 'curso', 'nivel',
            # y 'seccion', 'section'. También manejamos columnas combinadas como 'grado_seccion' o 'grado/seccion'.
            raw_grado = (r.get('grado') or r.get('grade') or r.get('grado_seccion') or r.get('grado/seccion') or r.get('curso') or r.get('nivel') or '')
            raw_seccion = (r.get('seccion') or r.get('section') or r.get('seccion_grado') or '')

            def _split_grado_seccion(text):
                if not text:
                    return ('', '')
                s = str(text).strip()
                # Normalize separators
                s2 = re.sub(r'[\|,;]+', '-', s)
                # Try pattern like '5-A' or '5 A' or '5/A'
                m = re.match(r"^\s*([A-Za-z0-9ñÑ°º]+)\s*[-/\\\s]+\s*([A-Za-z0-9ñÑ]+)\s*$", s2)
                if m:
                    return (m.group(1).strip(), m.group(2).strip())
                # Try reversed 'A-5' -> section-grade
                m2 = re.match(r"^\s*([A-Za-zñÑ]+)\s*[-/\\\s]+\s*([0-9]+)\s*$", s2)
                if m2:
                    return (m2.group(2).strip(), m2.group(1).strip())
                # If single token with space-separated tokens, take first as grado and last as seccion
                parts = s.split()
                if len(parts) >= 2:
                    return (parts[0].strip(), parts[-1].strip())
                # fallback: return text as grado and empty seccion
                return (s, '')

            grado_nombre, seccion_nombre = _split_grado_seccion(raw_grado)
            # if seccion separately provided, prefer that
            if raw_seccion:
                seccion_nombre = str(raw_seccion).strip()

            # Normalize strings
            if grado_nombre:
                grado_nombre = str(grado_nombre).strip()
            if seccion_nombre:
                seccion_nombre = str(seccion_nombre).strip()[:5]  # seccion max_length safety
            codigo_interno_val = (r.get('codigo_del_estudiante') or r.get('codigo_estudiante') or r.get('codigo') or r.get('codigo_interno') or '')
            estado_matricula_val = (r.get('estado_de_matricula') or r.get('estado_matricula') or r.get('matricula_estado') or '')
            observacion_val = (r.get('observacion') or r.get('observaciones') or r.get('obs') or '')

            ap_nombre = (r.get('apoderado_nombre') or r.get('tutor_nombre') or '').strip()
            ap_apellido = (r.get('apoderado_apellido') or r.get('tutor_apellido') or '').strip()
            ap_celular = (r.get('apoderado_celular') or r.get('tutor_celular') or '')
            ap_correo = (r.get('apoderado_correo') or r.get('tutor_correo') or '')
            if ap_nombre:
                ap_nombre = ap_nombre.strip()
            if ap_apellido:
                ap_apellido = ap_apellido.strip()
            if ap_correo:
                ap_correo = ap_correo.strip()

            if not dni or not nombre or not apellido:
                errors.append({**r, 'error': 'fila incompleta (dni/nombre/apellido)'} )
                self.stdout.write(self.style.WARNING(f'Se salta fila incompleta (dni/nombre/apellido): {r}'))
                continue

            # Grado / Seccion
            if grado_nombre:
                grado_obj, _ = Grado.objects.get_or_create(nombre=grado_nombre)
            else:
                grado_obj = None

            if seccion_nombre and grado_obj:
                seccion_obj, _ = Seccion.objects.get_or_create(nombre=seccion_nombre, grado=grado_obj)
            else:
                seccion_obj = None

            # Apoderado
            apoderado_obj = None
            if ap_correo:
                apoderado_obj, _ = Apoderado.objects.get_or_create(correo=ap_correo, defaults={
                    'nombre': ap_nombre or 'N/A',
                    'apellido': ap_apellido or 'N/A',
                    'celular': ap_celular or ''
                })
            elif ap_nombre or ap_apellido:
                apoderado_obj, _ = Apoderado.objects.get_or_create(nombre=ap_nombre, apellido=ap_apellido, defaults={'celular': ap_celular or '', 'correo': ''})

            # Crear o actualizar estudiante
            est_kwargs = {
                'nombre': nombre,
                'apellido': apellido,
                'dni': dni,
            }
            if fecha_nacimiento:
                est_kwargs['fecha_nacimiento'] = fecha_nacimiento
            if grado_obj:
                est_kwargs['grado'] = grado_obj
            if seccion_obj:
                est_kwargs['seccion'] = seccion_obj
            if apoderado_obj:
                est_kwargs['apoderado'] = apoderado_obj
            if periodo_override:
                est_kwargs['periodo'] = periodo_override
            # Campos opcionales nuevos
            if codigo_interno_val:
                est_kwargs['codigo_interno'] = str(codigo_interno_val).strip()
            if estado_matricula_val:
                est_kwargs['estado_matricula'] = str(estado_matricula_val).strip()
            if observacion_val:
                est_kwargs['observaciones'] = str(observacion_val).strip()

            estudiante = None
            try:
                estudiante = Estudiante.objects.get(dni=dni)
                # update
                for k, v in est_kwargs.items():
                    setattr(estudiante, k, v)
                estudiante.save()
                updated += 1
                created_flag = False
            except Estudiante.DoesNotExist:
                # Need grado and seccion; if missing, create placeholder
                if not est_kwargs.get('grado'):
                    grado_obj, _ = Grado.objects.get_or_create(nombre='Sin Grado')
                    est_kwargs['grado'] = grado_obj
                if not est_kwargs.get('seccion'):
                    # Seccion tiene max_length=5, usar placeholder corto
                    seccion_obj, _ = Seccion.objects.get_or_create(nombre='Sin', grado=est_kwargs['grado'])
                    est_kwargs['seccion'] = seccion_obj
                try:
                    estudiante = Estudiante.objects.create(**est_kwargs, codigo_qr=dni)
                except Exception as e:
                    errors.append({**r, 'error': f'error creando estudiante: {e}'})
                    self.stdout.write(self.style.ERROR(f'Error creando estudiante {dni}: {e}'))
                    continue
                created += 1
                created_flag = True

            # Generar imagen QR y guardar en MEDIA_ROOT/qrcodes/{dni}.png
            try:
                qr = qrcode.QRCode(version=1, box_size=10, border=4)
                qr.add_data(dni)
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                qr_path = os.path.join(media_qr_dir, f'{dni}.png')
                img.save(qr_path)
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Error generando QR para {dni}: {e}'))

            # update progress
            processed += 1
            if os.path.exists(status_path):
                _write_progress({'status': 'processing', 'started_at': _dt.datetime.now().isoformat(), 'processed': processed, 'total': total_rows})

        # Si hubo errores, escribir CSV de log en MEDIA_ROOT/import_logs
        if errors:
            log_dir = os.path.join(settings.MEDIA_ROOT or 'media', 'import_logs')
            os.makedirs(log_dir, exist_ok=True)
            timestamp = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
            log_path = os.path.join(log_dir, f'import_errors_{timestamp}.csv')
            # Obtener todas las keys
            keys = set()
            for e in errors:
                keys.update(e.keys())
            keys = list(keys)
            try:
                with open(log_path, 'w', newline='', encoding='utf-8') as lf:
                    writer = csv.DictWriter(lf, fieldnames=keys)
                    writer.writeheader()
                    for e in errors:
                        writer.writerow({k: (e.get(k) or '') for k in keys})
                self.stdout.write(self.style.WARNING(f'Errores de importación escritos en: {log_path}'))
                # If a status file exists next to the uploaded file, update it to point to the errors log
                try:
                    rel_log = os.path.join('import_logs', os.path.basename(log_path))
                    _write_progress({'status': 'processing', 'errors_log': rel_log, 'processed': processed, 'total': total_rows})
                except Exception:
                    pass
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'No se pudo escribir log de errores: {e}'))

        # Finalizar status file if present
        if os.path.exists(status_path):
            try:
                _write_progress({'status': 'done', 'finished_at': _dt.datetime.now().isoformat(), 'created': created, 'updated': updated})
            except Exception:
                pass

        self.stdout.write(self.style.SUCCESS(f'Importación finalizada. Creados: {created}, Actualizados: {updated}'))
